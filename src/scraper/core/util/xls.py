import csv
from io import BytesIO
from pathlib import Path

import openpyxl


def to_csv(xlsx_bytes: bytes, out_csv: Path, delimiter: str = ","):
    """
    Convert an in-memory XLSX (bytes) into a CSV file at out_csv.
    - Writes headers as the first row.
    - Replaces None cells with empty string for CSV output.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        writer.writerow(["" if h is None else h for h in headers_raw])
        for row in ws.iter_rows(min_row=2, values_only=True):
            writer.writerow(["" if v is None else v for v in row])
